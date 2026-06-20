import os
import sqlite3
from datetime import datetime
from aiogram import Bot, Dispatcher, F
from aiogram.types import Message, FSInputFile, InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
import asyncio

# Библиотеки для работы с Excel и изображениями
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenPyXLImage
from PIL import Image as PILImage

# ================= НАСТРОЙКИ БОТА =================
BOT_TOKEN = "8773740004:AAEEOkrBVK8oRjfdqQrp-yIzyIf_X7oWDNQ"
EXCEL_FILE = "report.xlsx"
IMG_DIR = "bot_images"
DB_FILE = "bot_data.db"

# Список ваших разделов (вкладок в Excel)
CATEGORIES = ["Фундамент", "Стены", "Крыша", "Электрика"]

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()

os.makedirs(IMG_DIR, exist_ok=True)

# Очередь состояний для пошагового опроса
class BotStates(StatesGroup):
    waiting_for_bti = State()

# ================= РАБОТА С БАЗОЙ ДАННЫХ =================
def init_db():
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                telegram_msg_id INTEGER,
                date_str TEXT,
                caption TEXT,
                category TEXT,
                bti_room TEXT,
                file_path TEXT
            )
        """)
        conn.commit()
        
        # Проверяем и добавляем новые колонки, если база старая
        cursor.execute("PRAGMA table_info(messages)")
        columns = [col[1] for col in cursor.fetchall()]
        if "telegram_msg_id" not in columns:
            cursor.execute("ALTER TABLE messages ADD COLUMN telegram_msg_id INTEGER")
            conn.commit()
        if "category" not in columns:
            cursor.execute("ALTER TABLE messages ADD COLUMN category TEXT")
            conn.commit()
        if "bti_room" not in columns:
            cursor.execute("ALTER TABLE messages ADD COLUMN bti_room TEXT")
            conn.commit()

def add_to_db(msg_id, date_str, caption, file_path):
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO messages (telegram_msg_id, date_str, caption, category, bti_room, file_path) VALUES (?, ?, ?, ?, ?, ?)",
            (msg_id, date_str, caption, "Разное", "Не указано", file_path)
        )
        conn.commit()

def update_caption_in_db(msg_id, new_caption):
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE messages SET caption = ? WHERE telegram_msg_id = ?", (new_caption, msg_id))
        conn.commit()
        return cursor.rowcount

def update_category_in_db(msg_id, category_name):
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE messages SET category = ? WHERE telegram_msg_id = ?", (category_name, msg_id))
        conn.commit()

def update_bti_in_db(msg_id, bti_text):
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE messages SET bti_room = ? WHERE telegram_msg_id = ?", (bti_text, msg_id))
        conn.commit()

def get_all_from_db():
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT date_str, bti_room, caption, category, file_path FROM messages")
        return cursor.fetchall()

def clear_db():
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM messages")
        conn.commit()


# ================= ГЕНЕРАЦИЯ EXCEL (БЕЗ КОЛОНКИ ДАТЫ) =================
def generate_excel(data_rows):
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    sheets = {}
    target_size = (150, 150)
    
    for date_str, bti_room, text, category, img_path in data_rows:
        sheet_name = category if category else "Разное"
        sheet_name = sheet_name[:30]
        
        if sheet_name not in sheets:
            ws = wb.create_sheet(title=sheet_name)
            # ТЕПЕРЬ ТУТ ТОЛЬКО ТРИ КОЛОНКИ: БТИ, Текст и Фото
            ws.append(["Помещение (БТИ)", "Текст (Подпись)", "Фотография"])
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 25
            sheets[sheet_name] = (ws, 2)
            
        ws, current_row = sheets[sheet_name]
        
        if os.path.exists(img_path):
            with PILImage.open(img_path) as img:
                img.thumbnail(target_size)
                img.save(img_path)
            
            # Записываем данные в новые позиции без даты
            ws.cell(row=current_row, column=1, value=bti_room)
            ws.cell(row=current_row, column=2, value=text)
            
            xl_img = OpenPyXLImage(img_path)
            ws.add_image(xl_img, f"C{current_row}") # Фотография вставляется в колонку C
            ws.row_dimensions[current_row].height = (target_size[1] * 0.75) + 10
        else:
            ws.cell(row=current_row, column=1, value=bti_room)
            ws.cell(row=current_row, column=2, value=f"{text} (Фото удалено)")
            
        sheets[sheet_name] = (ws, current_row + 1)

    if not wb.sheetnames:
        wb.create_sheet(title="Пусто")

    wb.save(EXCEL_FILE)


def get_categories_keyboard(msg_id):
    buttons = []
    for cat in CATEGORIES:
        buttons.append([InlineKeyboardButton(text=cat, callback_data=f"set_cat:{msg_id}:{cat}")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


# ================= ЛОГИКА ТЕЛЕГРАМА =================

# 1. Принимаем фото и выводим категории
@dp.message(F.photo)
async def handle_photo(message: Message):
    photo = message.photo[-1]
    caption_text = message.caption if message.caption else "Без описания"
    
    file_name = f"{photo.file_unique_id}.jpg"
    local_path = os.path.join(IMG_DIR, file_name)
    await bot.download(photo, destination=local_path)
    
    msg_date = message.date.replace(tzinfo=None)
    pretty_date = msg_date.strftime("%d.%m.%Y %H:%M")
    
    add_to_db(message.message_id, pretty_date, caption_text, local_path)
    
    await message.reply(
        text="📁 Выберите раздел для этого фото:",
        reply_markup=get_categories_keyboard(message.message_id)
    )


# 2. Ловим нажатие кнопки раздела и переключаемся на ожидание номера БТИ
@dp.callback_query(F.data.startswith("set_cat:"))
async def process_category_click(callback: CallbackQuery, state: FSMContext):
    _, msg_id_str, cat_name = callback.data.split(":", 2)
    msg_id = int(msg_id_str)
    
    update_category_in_db(msg_id, cat_name)
    
    # Запоминаем ID сообщения, чтобы связать с ним будущий текст БТИ
    await state.update_data(current_msg_id=msg_id)
    # Включаем режим ожидания текста номера БТИ
    await state.set_state(BotStates.waiting_for_bti)
    
    await callback.message.edit_text(
        text=f"✅ Раздел **{cat_name}** выбран.\n\n✏️ Теперь номер помещения по БТИ или отправьте прочерк `-`"
    )
    await callback.answer()


# 3. Ловим сам номер БТИ (принимает любой текст)
@dp.message(BotStates.waiting_for_bti)
async def process_bti_text(message: Message, state: FSMContext):
    bti_input = message.text.strip()
    
    user_data = await state.get_data()
    msg_id = user_data.get("current_msg_id")
    
    if msg_id:
        update_bti_in_db(msg_id, bti_input)
        await message.reply(f"📍 Помещение {bti_input} ")
    
    # Выходим из режима ожидания текста, бот снова готов к новым фото
    await state.clear()


# 4. Редактирование описания в реальном времени
@dp.edited_message()
async def handle_any_edited_message(message: Message):
    new_caption = message.caption if message.caption else "Без описания"
    updated_rows = update_caption_in_db(message.message_id, new_caption)
    if updated_rows > 0:
        print(f"✏️ [ПРАВКА #{message.message_id}] Текст изменен на: '{new_caption}'")


# 5. Команда /ping
@dp.message(Command("ping"))
async def check_status(message: Message):
    now = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
    await message.reply(f"🟢 **Я на связи!**\n🖥️ Компьютер включен.\n🕒 Время на ПК: {now}")


# 6. Команда /generate
@dp.message(Command("generate"))
async def send_report(message: Message):
    data_rows = get_all_from_db()
    if not data_rows:
        await message.reply("📭 База данных пуста!")
        return
        
    status_msg = await message.reply("⚙️ Формирую Excel-таблицу с номерами БТИ...")
    
    try:
        generate_excel(data_rows)
        
        excel_file = FSInputFile(EXCEL_FILE)
        await message.reply_document(excel_file, caption=f"🎉 Отчет готов! Все помещения зафиксированы. Всего фото: {len(data_rows)}")
        
        for _, _, _, _, img_path in data_rows:
            if os.path.exists(img_path):
                os.remove(img_path)
        if os.path.exists(EXCEL_FILE):
            os.remove(EXCEL_FILE)
            
        clear_db()
        print("🧹 База очищена.")
        
    except Exception as e:
        await message.reply(f"❌ Ошибка: {e}")
    finally:
        if status_msg:
            try:
                await status_msg.delete()
            except:
                pass


async def main():
    init_db()
    print("Бот запущен. Пошаговый опрос (Раздел -> БТИ) работает в штатном режиме!")
    await dp.start_polling(bot, skip_updates=False)

if __name__ == "__main__":
    asyncio.run(main())